"""
Stagebox Snapshots, Audit and Reports Routes (Pro Only)
"""

import json
import subprocess
import zipfile
import yaml
from datetime import datetime
from pathlib import Path
from typing import Optional, List
from flask import Blueprint, jsonify, request

from web import config
from web.config import VERSION
from web.edition import get_edition_name
from web.services import device_manager, is_building_active, get_active_building
from web.routes.pro.admin import require_admin, require_pro

bp = Blueprint('snapshots', __name__)

# Global paths
SNAPSHOT_SCRIPT = Path('/home/coredev/stagebox/shelly_snapshot.py')
INSTALLER_PROFILE_FILE = Path('/home/coredev/stagebox/data/installer_profile.json')
INSTALLER_LOGO_FILE = Path('/home/coredev/stagebox/data/installer_logo.png')


def _compare_device_configs(snapshot_dev: dict, live_dev: dict) -> List[str]:
    """Compare device configurations between snapshot and live state."""
    differences = []
    
    snap_config = snapshot_dev.get('config', {})
    live_config = live_dev.get('config', {})
    
    # Compare input settings
    for i in range(4):
        key = f'input:{i}'
        if key in snap_config or key in live_config:
            snap_input = snap_config.get(key, {})
            live_input = live_config.get(key, {})
            
            if snap_input.get('type') != live_input.get('type'):
                differences.append(f"{key}.type: {snap_input.get('type')} → {live_input.get('type')}")
            if snap_input.get('invert') != live_input.get('invert'):
                differences.append(f"{key}.invert: {snap_input.get('invert')} → {live_input.get('invert')}")
    
    # Compare switch settings
    for i in range(2):
        key = f'switch:{i}'
        if key in snap_config or key in live_config:
            snap_switch = snap_config.get(key, {})
            live_switch = live_config.get(key, {})
            
            if snap_switch.get('in_mode') != live_switch.get('in_mode'):
                differences.append(f"{key}.in_mode: {snap_switch.get('in_mode')} → {live_switch.get('in_mode')}")
            if snap_switch.get('initial_state') != live_switch.get('initial_state'):
                differences.append(f"{key}.initial_state: {snap_switch.get('initial_state')} → {live_switch.get('initial_state')}")
    
    # Compare cover settings
    for i in range(1):
        key = f'cover:{i}'
        if key in snap_config or key in live_config:
            snap_cover = snap_config.get(key, {})
            live_cover = live_config.get(key, {})
            
            for prop in ['in_mode', 'swap_inputs', 'invert_directions']:
                if snap_cover.get(prop) != live_cover.get(prop):
                    differences.append(f"{key}.{prop}: {snap_cover.get(prop)} → {live_cover.get(prop)}")
    
    # Compare sys.device.name
    snap_name = snap_config.get('sys', {}).get('device', {}).get('name')
    live_name = live_config.get('sys', {}).get('device', {}).get('name')
    if snap_name and live_name and snap_name != live_name:
        differences.append(f"sys.device.name: {snap_name} → {live_name}")
    
    # Compare webhooks count
    snap_hooks = len(snapshot_dev.get('webhooks', {}).get('hooks', []))
    live_hooks = len(live_dev.get('webhooks', {}).get('hooks', []))
    if snap_hooks != live_hooks:
        differences.append(f"webhooks: {snap_hooks} → {live_hooks}")
    
    # Compare KVS keys
    snap_kvs = set(snapshot_dev.get('kvs', {}).keys()) if snapshot_dev.get('kvs') else set()
    live_kvs = set(live_dev.get('kvs', {}).keys()) if live_dev.get('kvs') else set()
    if snap_kvs != live_kvs:
        added = live_kvs - snap_kvs
        removed = snap_kvs - live_kvs
        if added:
            differences.append(f"KVS added: {', '.join(added)}")
        if removed:
            differences.append(f"KVS removed: {', '.join(removed)}")
    
    return differences


def _create_snapshot_bundle(snapshots_dir: Path, snapshot_json_path: Path) -> Optional[Path]:
    """Create a ZIP bundle containing snapshot and all building files."""
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    zip_filename = f'snapshot_{timestamp}.zip'
    zip_path = snapshots_dir / zip_filename
    
    try:
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            # 1. Snapshot JSON (rename to snapshot.json in archive)
            if snapshot_json_path.exists():
                zf.write(snapshot_json_path, 'snapshot.json')
            
            # 2. Installer profile (global)
            if INSTALLER_PROFILE_FILE.exists():
                zf.write(INSTALLER_PROFILE_FILE, 'installer_profile.json')
            
            # 3. Installer logo (global, optional)
            if INSTALLER_LOGO_FILE.exists():
                zf.write(INSTALLER_LOGO_FILE, 'installer_logo.png')
            
            # 4. Building-specific files from DATA_DIR
            if config.DATA_DIR:
                # ip_state.json
                ip_state_file = config.DATA_DIR / 'ip_state.json'
                if ip_state_file.exists():
                    zf.write(ip_state_file, 'ip_state.json')
                
                # building_profile.json (optional)
                building_profile_file = config.DATA_DIR / 'building_profile.json'
                if building_profile_file.exists():
                    zf.write(building_profile_file, 'building_profile.json')
                
                # config.yaml
                config_file = config.DATA_DIR / 'config.yaml'
                if config_file.exists():
                    zf.write(config_file, 'config.yaml')
                
                # shelly_model_map.yaml (optional)
                model_map_file = config.DATA_DIR / 'shelly_model_map.yaml'
                if model_map_file.exists():
                    zf.write(model_map_file, 'shelly_model_map.yaml')
                
                # Scripts directory (optional)
                scripts_dir = config.DATA_DIR / 'scripts'
                if scripts_dir.exists() and scripts_dir.is_dir():
                    for script_file in scripts_dir.glob('*.js'):
                        zf.write(script_file, f'scripts/{script_file.name}')
        
        # Remove the temporary JSON file
        if snapshot_json_path.exists():
            snapshot_json_path.unlink()
        
        return zip_path
        
    except Exception as e:
        print(f"Error creating snapshot bundle: {e}")
        # Cleanup partial ZIP if it exists
        if zip_path.exists():
            zip_path.unlink()
        return None


def _extract_snapshot_json_from_zip(zip_path: Path) -> Optional[dict]:
    """Extract and parse snapshot.json from a ZIP bundle."""
    try:
        with zipfile.ZipFile(zip_path, 'r') as zf:
            if 'snapshot.json' in zf.namelist():
                with zf.open('snapshot.json') as f:
                    return json.load(f)
    except Exception as e:
        print(f"Error extracting snapshot from {zip_path}: {e}")
    return None


def _cleanup_old_snapshots(snapshots_dir: Path, keep: int = 10):
    """Keep only the N most recent snapshot ZIPs."""
    try:
        snapshots = sorted(snapshots_dir.glob('snapshot_*.zip'), reverse=True)
        for old_snap in snapshots[keep:]:
            old_snap.unlink()
    except Exception as e:
        print(f"Error cleaning up old snapshots: {e}")


# ===========================================================================
# Snapshots
# ===========================================================================

@bp.route('/api/building/snapshot', methods=['POST'])
@require_pro
def create_snapshot():
    """Create a new snapshot ZIP bundle of all devices and config files."""
    if not is_building_active():
        return jsonify({'success': False, 'error': 'No building active'}), 400
    
    snapshots_dir = config.DATA_DIR / 'snapshots'
    snapshots_dir.mkdir(parents=True, exist_ok=True)
    
    # Load config to get IP range
    config_file = config.DATA_DIR / 'config.yaml'
    if not config_file.exists():
        return jsonify({'success': False, 'error': 'Config file not found'}), 400
    
    try:
        with open(config_file, 'r', encoding='utf-8') as f:
            cfg = yaml.safe_load(f) or {}
    except Exception as e:
        return jsonify({'success': False, 'error': f'Cannot read config: {e}'}), 500
    
    # Get network settings from stage2.network
    stage2 = cfg.get('stage2', {})
    network = stage2.get('network', {})
    
    pool_start = network.get('pool_start')
    pool_end = network.get('pool_end')
    
    if not pool_start or not pool_end:
        return jsonify({'success': False, 'error': 'Network pool_start/pool_end not configured'}), 400
    
    # Check if snapshot script exists
    if not SNAPSHOT_SCRIPT.exists():
        return jsonify({'success': False, 'error': 'Snapshot script not found'}), 500
    
    try:
        # Run snapshot script
        result = subprocess.run(
            [
                'python3', str(SNAPSHOT_SCRIPT),
                '--ip-start', pool_start,
                '--ip-end', pool_end,
                '--output', str(snapshots_dir),
                '--timeout', '3',
                '--parallel', '20'
            ],
            capture_output=True,
            text=True,
            timeout=120  # 2 minutes max
        )
        
        if result.returncode != 0:
            return jsonify({
                'success': False,
                'error': result.stderr or 'Snapshot failed'
            }), 500
        
        # Find the newest snapshot JSON file created by the script
        json_snapshots = sorted(snapshots_dir.glob('shelly_snapshot_*.json'), reverse=True)
        if not json_snapshots:
            return jsonify({'success': False, 'error': 'No snapshot file created'}), 500
        
        latest_json = json_snapshots[0]
        
        # Load snapshot data for response
        with open(latest_json, 'r', encoding='utf-8') as f:
            snapshot_data = json.load(f)
        
        # Create ZIP bundle (this also removes the temp JSON)
        zip_path = _create_snapshot_bundle(snapshots_dir, latest_json)
        if not zip_path:
            return jsonify({'success': False, 'error': 'Failed to create snapshot bundle'}), 500
        
        # Cleanup old snapshots
        _cleanup_old_snapshots(snapshots_dir)
        
        # Also cleanup any leftover .json files
        for old_json in snapshots_dir.glob('shelly_snapshot_*.json'):
            try:
                old_json.unlink()
            except Exception:
                pass
        
        return jsonify({
            'success': True,
            'filename': zip_path.name,
            'timestamp': snapshot_data.get('snapshot_timestamp'),
            'device_count': len(snapshot_data.get('devices', [])),
            'summary': snapshot_data.get('summary', {})
        })
        
    except subprocess.TimeoutExpired:
        return jsonify({'success': False, 'error': 'Snapshot timed out (>2 minutes)'}), 500
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@bp.route('/api/building/snapshots', methods=['GET'])
@require_pro
def list_snapshots():
    """List all snapshots for the active building."""
    if not is_building_active():
        return jsonify({'success': False, 'error': 'No building active'}), 400
    
    try:
        snapshots_dir = config.DATA_DIR / 'snapshots'
        if not snapshots_dir.exists():
            return jsonify({'success': True, 'snapshots': [], 'has_recent': False})
        
        snapshots = []
        for zip_file in sorted(snapshots_dir.glob('snapshot_*.zip'), reverse=True):
            try:
                data = _extract_snapshot_json_from_zip(zip_file)
                if data:
                    # Handle both old format (snapshot_timestamp) and new format (created_at)
                    timestamp = data.get('snapshot_timestamp') or data.get('created_at')
                    # devices can be a list (from shelly_snapshot.py) or dict (from ip_state)
                    devices = data.get('devices', [])
                    device_count = len(devices) if isinstance(devices, list) else len(devices.keys())
                    
                    snapshots.append({
                        'filename': zip_file.name,
                        'timestamp': timestamp,
                        'device_count': device_count,
                        'scan_range': data.get('scan_range', ''),
                        'size_bytes': zip_file.stat().st_size
                    })
            except:
                pass
        
        # Check if there's a recent snapshot (within 7 days)
        has_recent = False
        if snapshots and snapshots[0].get('timestamp'):
            try:
                from datetime import timedelta
                latest_ts = datetime.fromisoformat(snapshots[0]['timestamp'].replace('Z', '+00:00'))
                # Make comparison timezone-naive
                if latest_ts.tzinfo:
                    latest_ts = latest_ts.replace(tzinfo=None)
                has_recent = (datetime.now() - latest_ts) < timedelta(days=7)
            except Exception:
                pass
        
        return jsonify({'success': True, 'snapshots': snapshots, 'has_recent': has_recent})
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@bp.route('/api/building/snapshot/<filename>', methods=['GET'])
@require_pro
def get_snapshot(filename):
    """Get details of a specific snapshot."""
    if not is_building_active():
        return jsonify({'success': False, 'error': 'No building active'}), 400
    
    try:
        snapshot_path = config.DATA_DIR / 'snapshots' / filename
        if not snapshot_path.exists():
            return jsonify({'success': False, 'error': 'Snapshot not found'}), 404
        
        # Handle both ZIP and JSON files
        if filename.endswith('.zip'):
            data = _extract_snapshot_json_from_zip(snapshot_path)
            if not data:
                return jsonify({'success': False, 'error': 'Could not read snapshot'}), 500
        else:
            with open(snapshot_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
        
        return jsonify({'success': True, 'snapshot': data})
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@bp.route('/api/building/snapshot/<filename>/download', methods=['GET'])
@require_pro
def download_snapshot(filename):
    """Download a snapshot ZIP file."""
    from flask import send_file
    
    if not is_building_active():
        return jsonify({'success': False, 'error': 'No building active'}), 400
    
    try:
        snapshot_path = config.DATA_DIR / 'snapshots' / filename
        if not snapshot_path.exists():
            return jsonify({'success': False, 'error': 'Snapshot not found'}), 404
        
        return send_file(
            str(snapshot_path),
            mimetype='application/zip',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@bp.route('/api/building/snapshot/<filename>', methods=['DELETE'])
@require_pro
def delete_snapshot(filename):
    """Delete a snapshot."""
    if not is_building_active():
        return jsonify({'success': False, 'error': 'No building active'}), 400
    
    try:
        snapshot_path = config.DATA_DIR / 'snapshots' / filename
        if not snapshot_path.exists():
            return jsonify({'success': False, 'error': 'Snapshot not found'}), 404
        
        snapshot_path.unlink()
        
        return jsonify({'success': True, 'message': 'Snapshot deleted'})
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ===========================================================================
# Audit
# ===========================================================================

@bp.route('/api/building/audit', methods=['POST'])
@require_pro
def run_audit():
    """Run audit: compare live state vs selected snapshot."""
    import tempfile
    
    if not is_building_active():
        return jsonify({'success': False, 'error': 'No building active'}), 400
    
    # Accept both 'snapshot' and 'snapshot_id' parameter names
    data = request.get_json(silent=True) or {}
    snapshot_id = data.get('snapshot') or data.get('snapshot_id')
    
    if not snapshot_id:
        return jsonify({'success': False, 'error': 'No snapshot selected'}), 400
    
    # Load config to get IP range
    config_file = config.DATA_DIR / 'config.yaml'
    if not config_file.exists():
        return jsonify({'success': False, 'error': 'Config file not found'}), 400
    
    try:
        with open(config_file, 'r', encoding='utf-8') as f:
            cfg = yaml.safe_load(f) or {}
    except Exception as e:
        return jsonify({'success': False, 'error': f'Cannot read config: {e}'}), 500
    
    # Get network settings from stage2.network
    stage2 = cfg.get('stage2', {})
    network = stage2.get('network', {})
    
    pool_start = network.get('pool_start')
    pool_end = network.get('pool_end')
    
    if not pool_start or not pool_end:
        return jsonify({'success': False, 'error': 'Network pool_start/pool_end not configured'}), 400
    
    # Check if snapshot script exists
    if not SNAPSHOT_SCRIPT.exists():
        return jsonify({'success': False, 'error': 'Snapshot script not found'}), 500
    
    # 1. Load reference snapshot from ZIP
    snapshots_dir = config.DATA_DIR / 'snapshots'
    snapshot_file = snapshots_dir / snapshot_id
    if not snapshot_file.exists():
        return jsonify({'success': False, 'error': f'Snapshot not found: {snapshot_id}'}), 404
    
    snapshot_json = _extract_snapshot_json_from_zip(snapshot_file)
    if not snapshot_json:
        return jsonify({'success': False, 'error': 'Cannot read snapshot'}), 500
    
    snapshot_timestamp = snapshot_json.get('snapshot_timestamp')
    snapshot_data = {}
    for dev in snapshot_json.get('devices', []):
        mac = dev.get('device_info', {}).get('mac', '').upper()
        if mac:
            snapshot_data[mac] = dev
    
    # 2. Run live scan (to temp file)
    with tempfile.TemporaryDirectory() as temp_dir:
        try:
            result = subprocess.run(
                [
                    'python3', str(SNAPSHOT_SCRIPT),
                    '--ip-start', pool_start,
                    '--ip-end', pool_end,
                    '--output', temp_dir,
                    '--timeout', '3',
                    '--parallel', '20'
                ],
                capture_output=True,
                text=True,
                timeout=120
            )
            
            if result.returncode != 0:
                return jsonify({
                    'success': False,
                    'error': result.stderr or 'Live scan failed'
                }), 500
            
            # Find the temp snapshot file
            temp_snapshots = list(Path(temp_dir).glob('shelly_snapshot_*.json'))
            if not temp_snapshots:
                return jsonify({'success': False, 'error': 'Live scan produced no data'}), 500
            
            with open(temp_snapshots[0], 'r', encoding='utf-8') as f:
                live_data = json.load(f)
                
        except subprocess.TimeoutExpired:
            return jsonify({'success': False, 'error': 'Live scan timeout'}), 500
    
    # 3. Build live_data dict keyed by MAC
    live_devices = {}
    for dev in live_data.get('devices', []):
        mac = dev.get('device_info', {}).get('mac', '').upper()
        if mac:
            live_devices[mac] = dev
    
    # 4. Compare snapshot vs live
    all_macs = set(snapshot_data.keys()) | set(live_devices.keys())
    
    audit_results = []
    summary = {'ok': 0, 'warning': 0, 'offline': 0, 'new': 0}
    
    for mac in sorted(all_macs):
        in_snapshot = mac in snapshot_data
        in_live = mac in live_devices
        
        audit_result = {
            'mac': mac,
            'in_snapshot': in_snapshot,
            'in_live': in_live,
            'differences': [],
            'status': 'ok'
        }
        
        # Get device info from available sources
        if in_live:
            live_dev = live_devices[mac]
            audit_result['ip'] = live_dev.get('ip')
            audit_result['name'] = live_dev.get('device_info', {}).get('name')
            audit_result['model'] = live_dev.get('device_info', {}).get('app') or live_dev.get('device_info', {}).get('model')
            audit_result['fw'] = live_dev.get('device_info', {}).get('ver')
        elif in_snapshot:
            snap_dev = snapshot_data[mac]
            audit_result['ip'] = snap_dev.get('ip')
            audit_result['name'] = snap_dev.get('device_info', {}).get('name')
            audit_result['model'] = snap_dev.get('device_info', {}).get('app') or snap_dev.get('device_info', {}).get('model')
            audit_result['fw'] = snap_dev.get('device_info', {}).get('ver')
        
        # Determine status and find differences
        if in_snapshot and not in_live:
            audit_result['status'] = 'offline'
            audit_result['differences'].append('Device offline (was in snapshot but not responding)')
            summary['offline'] += 1
        elif in_live and not in_snapshot:
            audit_result['status'] = 'new'
            audit_result['differences'].append('New device (not in reference snapshot)')
            summary['new'] += 1
        elif in_live and in_snapshot:
            # Compare live vs snapshot
            live_dev = live_devices[mac]
            snap_dev = snapshot_data[mac]
            
            # IP comparison
            live_ip = live_dev.get('ip')
            snap_ip = snap_dev.get('ip')
            if live_ip and snap_ip and live_ip != snap_ip:
                audit_result['differences'].append(f"IP changed: {snap_ip} → {live_ip}")
            
            # Name comparison
            live_name = live_dev.get('device_info', {}).get('name')
            snap_name = snap_dev.get('device_info', {}).get('name')
            if live_name and snap_name and live_name != snap_name:
                audit_result['differences'].append(f"Name changed: {snap_name} → {live_name}")
            
            # FW comparison
            live_fw = live_dev.get('device_info', {}).get('ver')
            snap_fw = snap_dev.get('device_info', {}).get('ver')
            if live_fw and snap_fw and live_fw != snap_fw:
                audit_result['differences'].append(f"Firmware updated: {snap_fw} → {live_fw}")
            
            # Deep config comparison
            config_diffs = _compare_device_configs(snap_dev, live_dev)
            for diff in config_diffs:
                audit_result['differences'].append(f"Config: {diff}")
            
            if audit_result['differences']:
                audit_result['status'] = 'warning'
                summary['warning'] += 1
            else:
                summary['ok'] += 1
        
        audit_results.append(audit_result)
    
    return jsonify({
        'success': True,
        'timestamp': live_data.get('snapshot_timestamp'),
        'snapshot_timestamp': snapshot_timestamp,
        'summary': summary,
        'devices': audit_results
    })


# ===========================================================================
# Reports
# ===========================================================================

REPORT_TRANSLATIONS = {
    'de': {
        'report_title': 'Installationsbericht',
        'generated': 'Erstellt',
        'snapshot_date': 'Snapshot vom',
        'customer': 'Kunde',
        'address': 'Adresse',
        'phone': 'Telefon',
        'email': 'E-Mail',
        'summary': 'Zusammenfassung',
        'devices': 'Geräte',
        'rooms': 'Räume',
        'device_types': 'Gerätetypen',
        'device_overview': 'Geräteübersicht',
        'name': 'Name',
        'room': 'Raum',
        'location': 'Ort',
        'model': 'Modell',
        'firmware': 'Firmware',
        'notes': 'Notizen',
        'legend': 'Legende',
        'legend_scripts': 'Scripts',
        'legend_webhooks': 'Webhooks',
        'legend_timers': 'Auto-Timer',
        'legend_schedules': 'Zeitpläne',
        'legend_kvs': 'Key-Value Store',
        'snapshot_reference': 'Snapshot-Referenz',
        'generated_with': 'Erstellt mit',
        'page': 'Seite',
        'of': 'von'
    },
    'en': {
        'report_title': 'Installation Report',
        'generated': 'Generated',
        'snapshot_date': 'Snapshot from',
        'customer': 'Customer',
        'address': 'Address',
        'phone': 'Phone',
        'email': 'Email',
        'summary': 'Summary',
        'devices': 'Devices',
        'rooms': 'Rooms',
        'device_types': 'Device Types',
        'device_overview': 'Device Overview',
        'name': 'Name',
        'room': 'Room',
        'location': 'Location',
        'model': 'Model',
        'firmware': 'Firmware',
        'notes': 'Notes',
        'legend': 'Legend',
        'legend_scripts': 'Scripts',
        'legend_webhooks': 'Webhooks',
        'legend_timers': 'Auto-Timers',
        'legend_schedules': 'Schedules',
        'legend_kvs': 'Key-Value Store',
        'snapshot_reference': 'Snapshot Reference',
        'generated_with': 'Generated with',
        'page': 'Page',
        'of': 'of'
    },
    'fr': {
        'report_title': 'Rapport d\'installation',
        'generated': 'Généré',
        'snapshot_date': 'Snapshot du',
        'customer': 'Client',
        'address': 'Adresse',
        'phone': 'Téléphone',
        'email': 'E-mail',
        'summary': 'Résumé',
        'devices': 'Appareils',
        'rooms': 'Pièces',
        'device_types': 'Types d\'appareils',
        'device_overview': 'Aperçu des appareils',
        'name': 'Nom',
        'room': 'Pièce',
        'location': 'Emplacement',
        'model': 'Modèle',
        'firmware': 'Firmware',
        'notes': 'Notes',
        'legend': 'Légende',
        'legend_scripts': 'Scripts',
        'legend_webhooks': 'Webhooks',
        'legend_timers': 'Minuteries auto',
        'legend_schedules': 'Programmations',
        'legend_kvs': 'Key-Value Store',
        'snapshot_reference': 'Référence snapshot',
        'generated_with': 'Généré avec',
        'page': 'Page',
        'of': 'sur'
    },
    'it': {
        'report_title': 'Rapporto di installazione',
        'generated': 'Generato',
        'snapshot_date': 'Snapshot del',
        'customer': 'Cliente',
        'address': 'Indirizzo',
        'phone': 'Telefono',
        'email': 'E-mail',
        'summary': 'Riepilogo',
        'devices': 'Dispositivi',
        'rooms': 'Stanze',
        'device_types': 'Tipi di dispositivi',
        'device_overview': 'Panoramica dispositivi',
        'name': 'Nome',
        'room': 'Stanza',
        'location': 'Posizione',
        'model': 'Modello',
        'firmware': 'Firmware',
        'notes': 'Note',
        'legend': 'Legenda',
        'legend_scripts': 'Script',
        'legend_webhooks': 'Webhook',
        'legend_timers': 'Timer automatici',
        'legend_schedules': 'Programmazioni',
        'legend_kvs': 'Key-Value Store',
        'snapshot_reference': 'Riferimento snapshot',
        'generated_with': 'Generato con',
        'page': 'Pagina',
        'of': 'di'
    },
    'nl': {
        'report_title': 'Installatierapport',
        'generated': 'Gegenereerd',
        'snapshot_date': 'Snapshot van',
        'customer': 'Klant',
        'address': 'Adres',
        'phone': 'Telefoon',
        'email': 'E-mail',
        'summary': 'Samenvatting',
        'devices': 'Apparaten',
        'rooms': 'Kamers',
        'device_types': 'Apparaattypen',
        'device_overview': 'Apparaatoverzicht',
        'name': 'Naam',
        'room': 'Kamer',
        'location': 'Locatie',
        'model': 'Model',
        'firmware': 'Firmware',
        'notes': 'Notities',
        'legend': 'Legenda',
        'legend_scripts': 'Scripts',
        'legend_webhooks': 'Webhooks',
        'legend_timers': 'Auto-timers',
        'legend_schedules': 'Schema\'s',
        'legend_kvs': 'Key-Value Store',
        'snapshot_reference': 'Snapshot referentie',
        'generated_with': 'Gegenereerd met',
        'page': 'Pagina',
        'of': 'van'
    }
}


def _generate_qr_code_base64(data: str) -> str:
    """Generate a QR code as base64 data URI."""
    try:
        import qrcode
        import io
        import base64
        
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=4,
            border=1,
        )
        qr.add_data(data)
        qr.make(fit=True)
        
        img = qr.make_image(fill_color="black", back_color="white")
        
        buffer = io.BytesIO()
        img.save(buffer, format='PNG')
        buffer.seek(0)
        
        b64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
        return f"data:image/png;base64,{b64}"
    except Exception as e:
        print(f"QR code generation failed: {e}")
        return ""


def _get_logo_base64() -> str:
    """Get installer logo as base64 data URI."""
    if not INSTALLER_LOGO_FILE.exists():
        return ""
    
    try:
        import base64
        with open(INSTALLER_LOGO_FILE, 'rb') as f:
            b64 = base64.b64encode(f.read()).decode('utf-8')
        return f"data:image/png;base64,{b64}"
    except Exception as e:
        print(f"Logo loading failed: {e}")
        return ""


def _load_model_mapping() -> dict:
    """Load model name mapping from YAML file."""
    model_map_file = config.DATA_DIR / 'shelly_model_map.yaml'
    if model_map_file.exists():
        try:
            with open(model_map_file, 'r', encoding='utf-8') as f:
                return yaml.safe_load(f) or {}
        except:
            pass
    return {}


def _generate_excel_report(devices, building_profile, installer_profile, t, safe_name, report_date, snapshot_date):
    """Generate Excel report with device overview."""
    from flask import send_file
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import io
    
    wb = Workbook()
    
    # Colors
    header_fill = PatternFill(start_color="2A5298", end_color="2A5298", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=14, color="2A5298")
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )
    
    ws = wb.active
    sheet_name = t.get('device_overview', 'Device Overview')[:31]
    ws.title = sheet_name
    
    # Header info
    ws['A1'] = t.get('report_title', 'Installation Report')
    ws['A1'].font = title_font
    ws.merge_cells('A1:L1')
    
    ws['A2'] = f"{t.get('generated', 'Generated')}: {report_date}"
    ws['A3'] = f"{t.get('snapshot_date', 'Snapshot')}: {snapshot_date}"
    
    # Object info
    row = 5
    if building_profile.get('object_name'):
        ws[f'A{row}'] = building_profile['object_name']
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
    if building_profile.get('customer_name'):
        ws[f'A{row}'] = f"{t.get('customer', 'Customer')}: {building_profile['customer_name']}"
        row += 1
    if building_profile.get('address'):
        ws[f'A{row}'] = f"{t.get('address', 'Address')}: {building_profile['address']}"
        row += 1
    
    row += 2
    
    # Table headers
    headers = [
        t.get('room', 'Room'),
        t.get('location', 'Location'),
        t.get('name', 'Name'),
        t.get('model', 'Model'),
        'IP',
        t.get('firmware', 'FW'),
        'MAC',
        'S', 'W', 'T', 'A', 'K'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='left')
        cell.border = thin_border
    
    # Data rows
    for device in devices:
        row += 1
        values = [
            device.get('room') or '-',
            device.get('location') or '-',
            device.get('friendly_name') or device.get('hostname') or '-',
            device.get('model_display') or device.get('model') or '-',
            device.get('ip') or '-',
            device.get('fw') or '-',
            f"...{device.get('mac_short', '')}" if device.get('mac_short') else '-',
            '✓' if device.get('has_scripts') else '',
            '✓' if device.get('has_webhooks') else '',
            '✓' if device.get('has_auto_timer') else '',
            '✓' if device.get('has_schedules') else '',
            '✓' if device.get('has_kvs') else ''
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left')
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 8
    ws.column_dimensions['G'].width = 10
    for col in 'HIJKL':
        ws.column_dimensions[col].width = 4
    
    # Save to buffer
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    filename = f"Report_{safe_name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return send_file(
        excel_buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@bp.route('/api/building/report', methods=['POST'])
@require_pro
def generate_report():
    """Generate installation report as PDF or Excel."""
    from flask import send_file, render_template
    import io
    
    if not is_building_active():
        return jsonify({'success': False, 'error': 'No building active'}), 400
    
    data = request.get_json(silent=True) or {}
    snapshot_id = data.get('snapshot_id', 'new')
    lang = data.get('language', 'de')
    output_format = data.get('format', 'pdf')
    
    if lang not in REPORT_TRANSLATIONS:
        lang = 'en'
    
    t = REPORT_TRANSLATIONS[lang]
    
    try:
        # 1. Get or create snapshot
        snapshots_dir = config.DATA_DIR / 'snapshots'
        snapshots_dir.mkdir(parents=True, exist_ok=True)
        snapshot_data = None
        snapshot_filename = ''
        
        if snapshot_id == 'new':
            # Load config to get IP range
            config_file = config.DATA_DIR / 'config.yaml'
            with open(config_file, 'r', encoding='utf-8') as f:
                cfg = yaml.safe_load(f) or {}
            
            stage2_net = cfg.get('stage2', {}).get('network', {})
            ip_start = stage2_net.get('pool_start')
            ip_end = stage2_net.get('pool_end')
            
            if not ip_start or not ip_end:
                return jsonify({'success': False, 'error': 'IP range not configured'}), 400
            
            # Create new snapshot
            result = subprocess.run(
                ['python3', str(SNAPSHOT_SCRIPT), '--ip-start', ip_start, '--ip-end', ip_end, '--output', str(snapshots_dir)],
                capture_output=True, text=True, timeout=120
            )
            if result.returncode != 0:
                return jsonify({'success': False, 'error': f'Snapshot failed: {result.stderr}'}), 500
            
            json_snapshots = sorted(snapshots_dir.glob('shelly_snapshot_*.json'), reverse=True)
            if not json_snapshots:
                return jsonify({'success': False, 'error': 'No snapshot created'}), 500
            
            latest_json = json_snapshots[0]
            with open(latest_json, 'r', encoding='utf-8') as f:
                snapshot_data = json.load(f)
            
            zip_path = _create_snapshot_bundle(snapshots_dir, latest_json)
            snapshot_filename = zip_path.name if zip_path else 'temp'
        else:
            snapshot_file = snapshots_dir / snapshot_id
            if not snapshot_file.exists():
                return jsonify({'success': False, 'error': f'Snapshot not found: {snapshot_id}'}), 404
            
            snapshot_data = _extract_snapshot_json_from_zip(snapshot_file)
            if not snapshot_data:
                return jsonify({'success': False, 'error': 'Cannot read snapshot'}), 500
            snapshot_filename = snapshot_id
        
        # 2. Extract device data
        snapshot_devices = snapshot_data.get('devices', [])
        snapshot_timestamp = snapshot_data.get('snapshot_timestamp', '')
        
        # 3. Load ip_state for room/location
        ip_state = {}
        ip_state_file = config.DATA_DIR / 'ip_state.json'
        if ip_state_file.exists():
            with open(ip_state_file, 'r', encoding='utf-8') as f:
                ip_state = json.load(f).get('devices', {})
        
        # 4. Load profiles
        installer_profile = {}
        if INSTALLER_PROFILE_FILE.exists():
            with open(INSTALLER_PROFILE_FILE, 'r', encoding='utf-8') as f:
                installer_profile = json.load(f)
        
        building_profile = {}
        building_profile_file = config.DATA_DIR / 'building_profile.json'
        if building_profile_file.exists():
            with open(building_profile_file, 'r', encoding='utf-8') as f:
                building_profile = json.load(f)
        
        # 5. Load model mapping
        model_map = _load_model_mapping()
        
        # 6. Build device list
        devices = []
        rooms = set()
        
        for snap_dev in snapshot_devices:
            device_info = snap_dev.get('device_info', {})
            dev_config = snap_dev.get('config', {})
            
            ip = snap_dev.get('ip', '')
            mac = device_info.get('mac', '').upper().replace(':', '')
            state_dev = ip_state.get(mac, {})
            
            hw_model = device_info.get('model', '')
            app_name = device_info.get('app', '')
            model_display = model_map.get(hw_model, app_name or hw_model)
            
            scripts = [v.get('name', k) for k, v in dev_config.items() if k.startswith('script:') and isinstance(v, dict)]
            webhooks = snap_dev.get('webhooks', {}).get('hooks', [])
            schedules = snap_dev.get('schedules', {}).get('jobs', [])
            kvs = snap_dev.get('kvs', {})
            
            has_auto_timer = any(isinstance(v, dict) and (v.get('auto_on') or v.get('auto_off')) for v in dev_config.values())
            
            device = {
                'mac': mac,
                'mac_short': mac[-6:] if len(mac) >= 6 else mac,
                'ip': ip,
                'hostname': device_info.get('id', ''),
                'friendly_name': state_dev.get('friendly_name', device_info.get('name', '')),
                'room': state_dev.get('room', ''),
                'location': state_dev.get('location', ''),
                'model': hw_model,
                'model_display': model_display,
                'fw': device_info.get('ver', ''),
                'has_scripts': len(scripts) > 0,
                'has_webhooks': len(webhooks) > 0,
                'has_auto_timer': has_auto_timer,
                'has_schedules': len(schedules) > 0,
                'has_kvs': len(kvs) > 0 if isinstance(kvs, dict) else False
            }
            devices.append(device)
            if device['room']:
                rooms.add(device['room'])
        
        devices.sort(key=lambda d: (d.get('room') or 'zzz', d.get('friendly_name') or d.get('hostname') or 'zzz'))
        
        # 7. Format dates
        report_date = datetime.now().strftime('%d.%m.%Y %H:%M')
        try:
            snap_dt = datetime.fromisoformat(snapshot_timestamp.replace('Z', '+00:00'))
            snapshot_date = snap_dt.strftime('%d.%m.%Y %H:%M')
        except:
            snapshot_date = snapshot_timestamp
        
        object_name = building_profile.get('object_name', get_active_building())
        safe_name = "".join(c if c.isalnum() or c in ' -_' else '_' for c in object_name)
        
        # 8. Generate output
        if output_format == 'xlsx':
            return _generate_excel_report(devices, building_profile, installer_profile, t, safe_name, report_date, snapshot_date)
        else:
            # PDF
            for device in devices:
                device['qr_code'] = _generate_qr_code_base64(f"http://{device['ip']}") if device.get('ip') else ''
            
            html_content = render_template(
                'report_template.html',
                t=t,
                lang=lang,
                building_name=get_active_building(),
                installer_profile=installer_profile,
                installer_logo=_get_logo_base64(),
                building_profile=building_profile,
                devices=devices,
                rooms=rooms,
                model_count=len(set(d['model_display'] for d in devices)),
                report_date=report_date,
                snapshot_date=snapshot_date,
                snapshot_filename=snapshot_filename,
                version=VERSION,
                edition_name=get_edition_name()
            )
            
            from weasyprint import HTML
            
            pdf_buffer = io.BytesIO()
            # Use the web directory as base_url for CSS/images
            web_dir = Path(__file__).resolve().parent.parent.parent
            HTML(string=html_content, base_url=str(web_dir)).write_pdf(pdf_buffer)
            pdf_buffer.seek(0)
            
            filename = f"Report_{safe_name}_{datetime.now().strftime('%Y%m%d')}.pdf"
            
            return send_file(
                pdf_buffer,
                mimetype='application/pdf',
                as_attachment=True,
                download_name=filename
            )
    
    except ImportError as e:
        return jsonify({'success': False, 'error': f'Missing dependency: {e}. Install: pip install weasyprint qrcode openpyxl'}), 500
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500